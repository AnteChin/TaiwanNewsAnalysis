import jieba.analyse
import json
from openpyxl import Workbook
from collections import Counter
from banlist import banlist
from chinesetrans.langconv import *


def Traditional2Simplified(sentence):
    '''
    将sentence中的繁体字转为简体字
    :param sentence: 待转换的句子
    :return: 将句子中繁体字转换为简体字之后的句子
    '''
    sentence = Converter('zh-hans').convert(str(sentence))
    return sentence


def Simplified2Traditional(sentence):
    '''
    将sentence中的简体字转为繁体字
    :param sentence: 待转换的句子
    :return: 将句子中简体字转换为繁体字之后的句子
    '''
    sentence = Converter('zh-hant').convert(str(sentence))
    return sentence


with open('result.json', encoding='gbk') as f:
    j = f.read()
    dic = json.loads(j)  # 把json文件变成dict
    raw_contexts = []
    raw_titles = []
    for i in dic:
        raw_titles.append(i["title"])  # 读取dic文件中'name'的值
    for i in dic:
        raw_contexts.append(i["context"])  # 读取dic文件中'name'的值
    f.close()

contexts = ''
titles = ''
# 分析正文用这段
for each in raw_contexts:
    for i in each:
        contexts += i
        contexts += ' '
# 分析标题用这段
for i in raw_titles:
    titles += i
    titles += ' '  # 把所有的值串成一个string

st_word = Traditional2Simplified(titles)
sc_word = Traditional2Simplified(contexts)
print(st_word)
print(sc_word)
# 用jieba包分词
cut_titles = jieba.cut(st_word, False)
cut_contexts = jieba.cut(sc_word, False)
output_titles = []
output_contexts = []
# 使用禁用词
for i in cut_titles:
    if not isinstance(i, str):
        i = str(i)
    if i not in banlist:
        output_titles.append(i)
toptitles = Counter(output_titles).most_common(200)

for i in cut_contexts:
    if not isinstance(i, str):
        i = str(i)
    if i not in banlist:
        output_contexts.append(i)
topcontexts = Counter(output_contexts).most_common(400)

#  创建excel文件
workbook = Workbook()
worksheet1 = workbook.create_sheet("标题分词")
row = 2
worksheet1.cell(column=1, row=1, value="关键词")
worksheet1.cell(column=2, row=1, value="出现次数")
for top in toptitles:
    worksheet1.cell(column=1, row=row, value=Simplified2Traditional(top[0]))
    worksheet1.cell(column=2, row=row, value=Simplified2Traditional(top[1]))
    row = row + 1

worksheet2 = workbook.create_sheet("正文分词")
worksheet2.cell(column=1, row=1, value="关键词")
worksheet2.cell(column=2, row=1, value="出现次数")
row2 = 2
for top in topcontexts:
    worksheet2.cell(column=1, row=row2, value=Simplified2Traditional(top[0]))
    worksheet2.cell(column=2, row=row2, value=Simplified2Traditional(top[1]))
    row2 = row2 + 1

workbook.save("自由时报.xlsx")
