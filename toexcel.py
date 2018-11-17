import jieba.analyse
import json
from openpyxl import Workbook
from collections import Counter
from banlist import banlist

with open('result.json', encoding='gbk') as f:
    j = f.read()
    dic = json.loads(j)  # 把json文件变成dict

    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "原始数据"
    worksheet1.cell(column=1, row=1, value="标题")
    worksheet1.cell(column=2, row=1, value="日期")
    worksheet1.cell(column=3, row=1, value="正文")
    row = 2
    for i in dic:
        worksheet1.cell(column=1, row=row, value=i["title"])
        worksheet1.cell(column=2, row=row, value=i['date'])
        context = ''
        for c in i['context']:
            if c.endswith('...'):
                continue
            else:
                context = context + c
        worksheet1.cell(column=3, row=row, value=context)
        row = row + 1
    workbook.save("raw.xlsx")
    f.close()

